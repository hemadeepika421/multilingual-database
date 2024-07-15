from django.shortcuts import render,redirect
from django.http import HttpResponse
import os
import base64
from io import BytesIO
from PyPDF2 import PdfReader
import openpyxl
from django.conf import settings
from .models import sindhi,northeast,malayalam,contact,language,literary_work
from django.contrib.auth.models import User
from django.contrib.auth import login,logout,authenticate
from django.core.paginator import EmptyPage, Paginator, PageNotAnInteger
import matplotlib.pyplot as plt
from wordcloud import WordCloud

# sindhis=sindhi.objects.all()
# # print(type(sindhis.author_first_name))
# for obj in sindhis:
#     isbn=obj.accession_number
#     name=obj.title
#     auth=(obj.author_first_name )+(obj.author_last_name)
#     lan='Sindhi'
#     pub=obj.publisher
#     oth='Keyword: '+(obj.keyword)+'#Place: '+(obj.place)
#     lang=language.objects.filter(name='Sindhi')[0]
#     lit=literary_work.objects.create(name=name,isbn_no=isbn,author_name=auth,publisher=pub,language1=lan,others=oth,lang=lang)
#     lit.save()
#     print(1)

def home(request):
    if request.method=="POST":
        name=request.POST.get('name')
        number=request.POST.get('number')
        email=request.POST.get('email')
        subject=request.POST.get('subject')
        text=request.POST.get('text')
        mssg=contact.objects.create(name=name,number=number,email=email,subject=subject,text=text)
        mssg.save()
    return render(request,"citation/index1.html")

def index2(request):
    return render(request,"citation/index2.html")

def admin_page(request):
    return redirect('index')

def index(request):
    if request.user.is_superuser:
        context={
            'mssg':contact.objects.all(),
            'lang':language.objects.all()
        }
        if request.method=='POST':
            if 'newlang' in request.POST:
                file=request.FILES['excel_sheet']
                newlang=request.POST.get('newlang')
                title=int(request.POST.get('title'))
                author=int(request.POST.get('author'))
                isbn=int(request.POST.get('isbn'))
                publisher=int(request.POST.get('publisher'))
                langua=int(request.POST.get('language'))
                lan=language.objects.create(name=newlang)
                lan.save()
                nums=[title,author,isbn,publisher,langua]
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.worksheets[0]
                headings = [str(cell.value) for cell in sheet[1]]
                col=sheet.max_column
                for row in sheet.iter_rows(min_row=2,values_only=True):
                    t=str(row[title-1])
                    a=str(row[author-1])
                    s=str(row[isbn-1])
                    p=str(row[publisher-1])
                    if(langua==0):
                        l=newlang
                    else:
                        l=str(row[langua-1])
                    o=''
                    for i in range(1,col+1):
                        if i not in nums:
                            o+=headings[i-1]+': '+str(row[i-1])+'#'
                    lit=literary_work.objects.create(name=t,author_name=a,isbn_no=s,publisher=p,language1=l,others=o,lang=lan)
                    lit.save()
            else:
                file=request.FILES['excel_sheet']
                lang=request.POST.get('lang')
                title=int(request.POST.get('title'))
                author=int(request.POST.get('author'))
                isbn=int(request.POST.get('isbn'))
                publisher=int(request.POST.get('publisher'))
                langua=int(request.POST.get('language'))
                lan=language.objects.filter(name=lang)[0]
                nums=[title,author,isbn,publisher,langua]
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.worksheets[0]
                headings = [str(cell.value) for cell in sheet[1]]
                col=sheet.max_column
                for row in sheet.iter_rows(min_row=2,values_only=True):
                    t=str(row[title-1])
                    a=str(row[author-1])
                    s=str(row[isbn-1])
                    p=str(row[publisher-1])
                    if(langua==0):
                        l=lang
                    else:
                        l=str(row[langua-1])
                    o=''
                    for i in range(1,col+1):
                        if i not in nums:
                            o+=headings[i-1]+': '+str(row[i-1])+'#'
                    lit=literary_work.objects.create(name=t,author_name=a,isbn_no=s,publisher=p,language1=l,others=o,lang=lan)
                    lit.save()

        return render(request,"citation/index.html",context=context)
    else:
        return redirect(admin_page)

def contacts(request):
    if request.user.is_superuser:
        context={'mssg':contact.objects.all()}
        return render(request,'citation/cont.html',context)
    else:
        return redirect(admin_page)

def delcont(request,id):
    cont=contact.objects.filter(id=id)[0]
    cont.delete()
    return redirect('contacts')

def Sindhi(request):
    lang=language.objects.filter(name='Sindhi')[0]
    lang1=language.objects.filter(name='NorthEast')[0]
    obj=literary_work.objects.filter(lang=lang)
    obj1=literary_work.objects.filter(lang=lang1)
    obj2=malayalam.objects.all()
    lis=[len(obj),len(obj1),len(obj2)]
    context={"data":obj,
             "num0":lis[0],
             "num1":lis[1],
             "num2":lis[2]
            }
    # if request.method=="POST":
    #     searched=request.POST.get("searched")
    #     print(searched)
    #     # searche=searched.upper()
    #     context['data']=sindhi.objects.filter(title__icontains=searched)
    #     i=request.POST.get("type_of_search")
    #     j=int(i)
    #     if(j==1):
    #         context['data']=sindhi.objects.filter(title__icontains=searched)
    #     elif j==2:
    #         context['data']=sindhi.objects.filter(accession_number__icontains=searched)
    #     elif j==3:
    #         context['data']=sindhi.objects.filter(author_first_name__icontains=searched)
    #     elif j==4:
    #         context['data']=sindhi.objects.filter(author_last_name__icontains=searched)
    #     elif j==5:
    #         context['data']=sindhi.objects.filter(keyword__icontains=searched)
    #     elif j==6:
    #         context['data']=sindhi.objects.filter(publisher__icontains=searched)
    #     elif j==7:
    #         context['data']=sindhi.objects.filter(place__icontains=searched)
        
        # return render(request,"citation/sindhi.html",context=context)
        
    return render(request,"citation/sindhi.html",context=context)
def Northeast(request):
    obj=sindhi.objects.all()
    obj1=northeast.objects.all()
    obj2=malayalam.objects.all()
    lis=[len(obj),len(obj1),len(obj2)]
    context={"data":obj1,
             "num0":lis[0],
             "num1":lis[1],
             "num2":lis[2]
            }
    if request.method=="POST":
        searched=request.POST.get("search")
        print(searched)
        # searche=searched.upper()

        context['data']=northeast.objects.filter(title__icontains=searched)
        i=request.POST.get("type_of_search")
        j=int(i)
        if(j==1):
            context['data']=northeast.objects.filter(title__icontains=searched)
        elif j==2:
            context['data']=northeast.objects.filter(accession_number__icontains=searched)
        elif j==3:
            context['data']=northeast.objects.filter(class_number__icontains=searched)
        elif j==4:
            context['data']=northeast.objects.filter(author_name__icontains=searched)
        elif j==5:
            context['data']=northeast.objects.filter(year__icontains=searched)
        elif j==6:
            context['data']=northeast.objects.filter(publisher__icontains=searched)
        elif j==7:
            context['data']=northeast.objects.filter(collection__icontains=searched)
        elif j==8:
            context['data']=northeast.objects.filter(language__icontains=searched)
        return render(request,"citation/northeast.html",context=context)
    return render(request,"citation/northeast.html",context=context)
def Malayalam(request):
    obj=sindhi.objects.all()
    obj1=northeast.objects.all()
    obj2=malayalam.objects.all()
    lis=[len(obj),len(obj1),len(obj2)]
    context={"data":obj2,
             "num0":lis[0],
             "num1":lis[1],
             "num2":lis[2]
            }
    if request.method=="POST":
        searched=request.POST.get("search")
        # searche=searched.upper()

        context['data']=malayalam.objects.filter(malayalam_title__icontains=searched)
        i=request.POST.get("type_of_search")
        j=int(i)
        if(j==1):
            context['data']=malayalam.objects.filter(malayalam_title__icontains=searched)
        elif j==2:
            context['data']=malayalam.objects.filter(engish_title__icontains=searched)
        elif j==3:
            context['data']=malayalam.objects.filter(isbn_no__icontains=searched)
        elif j==4:
            context['data']=malayalam.objects.filter(author_name__icontains=searched)
        elif j==5:
            context['data']=malayalam.objects.filter(subject__icontains=searched)
        elif j==6:
            context['data']=malayalam.objects.filter(genre__icontains=searched)
        elif j==7:
            context['data']=malayalam.objects.filter(publisher__icontains=searched)
        elif j==8:
            context['data']=malayalam.objects.filter(year_of_dc__icontains=searched)
        elif j==9:
            context['data']=malayalam.objects.filter(year_of_pub__icontains=searched)
        return render(request,"citation/malayalam.html",context=context)
    return render(request,"citation/malayalam.html",context=context)

def sindhidesc(request,id):
    obj=literary_work.objects.filter(id=id)[0]
    st=obj.others
    lis=st.split('#')
    author = obj.author_name
    book_title = obj.name
    publication = obj.publisher
    isbn = obj.isbn_no
    keyword = lis[0]
    place = lis[1]

    citation = f"{author}. \"{book_title},\" {publication}, {keyword}, ISBN: {isbn}, {place}."
    context={
        'data':obj,
        'list':lis,
        'citation':citation 
    }
    return render(request,"citation/sindhidesc.html",context)

def generate_citation(data):
    citation = f"<h1>{data['title']}</h1>\n"
    citation += "<div class='book-info'>\n"
    citation += f"<p><strong>Author:</strong> {data['author_name']}</p>\n"
    citation += f"<p><strong>Publication:</strong> {data['publisher']}</p>\n"
    citation += f"<p><strong>Accession number:</strong> {data['accession_number']}</p>\n"
    citation += f"<p><strong>Class number:</strong> {data['class_number']}</p>\n"
    citation += f"<p><strong>Year:</strong> {data['year']}</p>\n"
    citation += f"<p><strong>Collection:</strong> {data['collection']}</p>\n"
    citation += f"<p><strong>Language:</strong> {data['language']}</p>\n"
    citation += "<p><strong>Open Access:</strong> No</p>\n"
    citation += "</div>"
    return citation

def northeastdesc(request,id):
    obj=northeast.objects.filter(id=id)[0]
    citation = generate_citation(obj)
    print(citation)
    context={
        'data':obj,
        "citation":citation 
    }
    return render(request,"citation/northeastdesc.html",context)

def malayalamdesc(request,id):
    obj=malayalam.objects.filter(id=id)[0]
    context={
        'data':obj
    }
    return render(request,"citation/malayalamdesc.html",context)

def main_page(request):
    lang=language.objects.all()
    lan=language.objects.all()[0]
    for i in lang:
        i.num=i.literary_work_set.count()
    if request.method=='POST':
        search=request.POST.get('searched')
        d1=literary_work.objects.filter(lang=lan,name__icontains=search)
        d2=literary_work.objects.filter(lang=lan,isbn_no__icontains=search)
        d3=literary_work.objects.filter(lang=lan,author_name__icontains=search)
        d4=literary_work.objects.filter(lang=lan,publisher__icontains=search)
        d5=literary_work.objects.filter(lang=lan,language1__icontains=search)
        data1=d1.union(d2,d3,d4,d5)
        data=data1.order_by('id')
    else:
        data=literary_work.objects.filter(lang=lan)
    pagin=Paginator(data,per_page=100)
    page_number = request.GET.get('page')  # Get the current page number
    try:
        mod_page = pagin.page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        mod_page = pagin.page(1)
    except EmptyPage:
        # If page is out of range (e.g., 9999), deliver last page of results
        mod_page = pagin.page(pagin.num_pages)
    return render(request,'citation/sindhi.html',context={'lang':lang,'data':mod_page})

def main_page_id(request,id):
    lang=language.objects.all()
    lan=language.objects.filter(id=id)[0]
    for i in lang:
        i.num=i.literary_work_set.count()
    if request.method=='POST':
        search=request.POST.get('searched')
        d1=literary_work.objects.filter(lang=lan,name__icontains=search)
        d2=literary_work.objects.filter(lang=lan,isbn_no__icontains=search)
        d3=literary_work.objects.filter(lang=lan,author_name__icontains=search)
        d4=literary_work.objects.filter(lang=lan,publisher__icontains=search)
        d5=literary_work.objects.filter(lang=lan,language1__icontains=search)
        data1=d1.union(d2,d3,d4,d5)
        data=data1.order_by('id')
    else:
        data=literary_work.objects.filter(lang=lan)
    pagin=Paginator(data,per_page=100)
    page_number = request.GET.get('page')  # Get the current page number
    try:
        mod_page = pagin.page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        mod_page = pagin.page(1)
    except EmptyPage:
        # If page is out of range (e.g., 9999), deliver last page of results
        mod_page = pagin.page(pagin.num_pages)
    return render(request,'citation/sindhi.html',context={'lang':lang,'data':mod_page})  

from io import BytesIO
from django.shortcuts import render
from PyPDF2 import PdfReader
from wordcloud import WordCloud

def textextr(request):
    text = ""
    paragraphs = []  # List to hold paragraphs of text
    wordcloud_base64 = None  # Base64 string for the WordCloud image

    if request.method == 'POST':
        file = request.FILES.get('pdfFile')
        if file:
            pdf_data = file.read()
            pdf_file = BytesIO(pdf_data)
            reader = PdfReader(pdf_file)
            for page_number in range(len(reader.pages)):
                page = reader.pages[page_number]
                page_text = page.extract_text()
                text += page_text
                paragraphs.extend(page_text.split('\n\n'))  # Split text into paragraphs
                
            # Create a word cloud object
            wordcloud = WordCloud()
            wordcloud.generate(text)

            # Convert WordCloud image to base64 string
            img_buffer = BytesIO()
            wordcloud.to_image().save(img_buffer, format="PNG")
            wordcloud_base64 = base64.b64encode(img_buffer.getvalue()).decode()

    return render(request, 'citation/text.html', {'paragraphs': paragraphs, 'wordcloud_base64': wordcloud_base64})
    
    
